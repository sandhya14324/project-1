"""
Builds a DASHBOARD sheet on student_analysis_results.xlsx:
- 4 charts (subject averages, pass/fail pie, gender comparison, test prep impact)
- summary number boxes at the top
Run this AFTER analysis.py has created student_analysis_results.xlsx.
"""

import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FILE = "student_analysis_results.xlsx"
wb = openpyxl.load_workbook(FILE)

all_data = wb["All Data"]
stats = wb["Subject Stats"]
gender = wb["Gender Comparison"]
prep = wb["Test Prep Impact"]

if "DASHBOARD" in wb.sheetnames:
    del wb["DASHBOARD"]
dash = wb.create_sheet("DASHBOARD", 0)

# ---- helper sheet with small chart-source tables (kept off to the side) ----
src = wb.create_sheet("chart_source")

# Subject averages
src["A1"] = "Subject"
src["B1"] = "Average"
subjects = ["math score", "reading score", "writing score"]
for i, s in enumerate(subjects, start=2):
    row = [r for r in stats.iter_rows(min_row=2, values_only=True) if r[0] == s][0]
    src.cell(row=i, column=1, value=s.title())
    src.cell(row=i, column=2, value=round(row[1], 2))

# Pass/Fail for math (score >= 40 = pass)
math_scores = [c[0] for c in all_data.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True)]
passed = sum(1 for m in math_scores if m >= 40)
failed = len(math_scores) - passed
src["D1"] = "Result"
src["E1"] = "Count"
src["D2"] = "Pass"
src["E2"] = passed
src["D3"] = "Fail"
src["E3"] = failed

# Gender comparison
src["G1"] = "Gender"
src["H1"] = "Math"
src["I1"] = "Reading"
src["J1"] = "Writing"
r = 2
for row in gender.iter_rows(min_row=2, values_only=True):
    src.cell(row=r, column=7, value=row[0].title())
    src.cell(row=r, column=8, value=round(row[1], 2))
    src.cell(row=r, column=9, value=round(row[2], 2))
    src.cell(row=r, column=10, value=round(row[3], 2))
    r += 1

# Test prep impact
src["L1"] = "Test Prep"
src["M1"] = "Avg Math"
src["N1"] = "Avg Reading"
r = 2
for row in prep.iter_rows(min_row=2, values_only=True):
    src.cell(row=r, column=12, value=row[0].title())
    src.cell(row=r, column=13, value=round(row[1], 2))
    src.cell(row=r, column=14, value=round(row[2], 2))
    r += 1

# ---- Summary boxes ----
dash["A1"] = "STUDENT PERFORMANCE DASHBOARD"
dash["A1"].font = Font(size=16, bold=True, color="FFFFFF")
dash.merge_cells("A1:H1")
dash["A1"].fill = PatternFill("solid", fgColor="1F3864")
dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
dash.row_dimensions[1].height = 28

total_students = all_data.max_row - 1
overall_pass = round(passed / total_students * 100, 1)
highest = max(c[0] for c in all_data.iter_rows(min_row=2, values_only=True) if False) if False else None
totals_col = [row[-2] for row in all_data.iter_rows(min_row=2, values_only=True)]  # total_score col
highest_score = max(totals_col)
lowest_subject = min(zip(["Math", "Reading", "Writing"],
                          [src["B2"].value, src["B3"].value, src["B4"].value]),
                      key=lambda x: x[1])[0]

boxes = [
    ("Total Students", total_students, "2E75B6"),
    ("Overall Pass Rate", f"{overall_pass}%", "548235"),
    ("Highest Score", highest_score, "BF8F00"),
    ("Lowest Avg Subject", lowest_subject, "C00000"),
]

col = 1
for label, value, color in boxes:
    c1 = dash.cell(row=3, column=col, value=label)
    c1.font = Font(bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=color)
    c1.alignment = Alignment(horizontal="center")
    c2 = dash.cell(row=4, column=col, value=value)
    c2.font = Font(bold=True, size=14, color=color)
    c2.alignment = Alignment(horizontal="center")
    dash.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
    dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    col += 2

for c in range(1, 9):
    dash.column_dimensions[get_column_letter(c)].width = 14

# ---- Chart 1: Bar - Subject vs Average Score ----
chart1 = BarChart()
chart1.title = "Average Score by Subject"
chart1.y_axis.title = "Average Score"
chart1.x_axis.title = "Subject"
data = Reference(src, min_col=2, min_row=1, max_row=4)
cats = Reference(src, min_col=1, min_row=2, max_row=4)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
dash.add_chart(chart1, "A6")

# ---- Chart 2: Pie - Pass vs Fail (Math) ----
chart2 = PieChart()
chart2.title = "Pass vs Fail — Math"
data = Reference(src, min_col=5, min_row=1, max_row=3)
cats = Reference(src, min_col=4, min_row=2, max_row=3)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
dash.add_chart(chart2, "F6")

# ---- Chart 3: Clustered Bar - Male vs Female avg score per subject ----
chart3 = BarChart()
chart3.type = "col"
chart3.grouping = "clustered"
chart3.title = "Male vs Female — Average Score per Subject"
chart3.y_axis.title = "Average Score"
data = Reference(src, min_col=8, max_col=10, min_row=1, max_row=3)
cats = Reference(src, min_col=7, min_row=2, max_row=3)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
dash.add_chart(chart3, "A22")

# ---- Chart 4: Bar - Test prep completed vs not, avg scores ----
chart4 = BarChart()
chart4.type = "col"
chart4.grouping = "clustered"
chart4.title = "Test Prep Completed vs Not — Average Scores"
chart4.y_axis.title = "Average Score"
data = Reference(src, min_col=13, max_col=14, min_row=1, max_row=3)
cats = Reference(src, min_col=12, min_row=2, max_row=3)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
dash.add_chart(chart4, "F22")

wb.save(FILE)
print("DASHBOARD sheet created with 4 charts and summary boxes.")